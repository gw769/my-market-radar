from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.marketplace import AnalysisRun, ListingSnapshot

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RESULT_STATUSES = ("completed", "partial")
FORMULA_PREFIXES = ("=", "+", "-", "@")


def _excel_safe(value):
    """Keep untrusted marketplace/user strings as literal Excel text, never formulas."""
    if not isinstance(value, str):
        return value
    candidate = value.lstrip()
    if candidate.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def _append(ws, values) -> None:
    ws.append([_excel_safe(value) for value in values])


def _dash(value):
    return "—" if value is None or value == "" else value


def _header(ws, labels: list[str]) -> None:
    _append(ws, labels)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_report(db: Session, run: AnalysisRun) -> BytesIO:
    keyword = run.tracked_keyword
    rows = db.query(ListingSnapshot).filter(ListingSnapshot.run_id == run.id).order_by(ListingSnapshot.platform, ListingSnapshot.search_rank).all()
    wb = Workbook()
    summary = wb.active
    summary.title = "综合结论"
    _append(summary, ["MY Marketplace 公开竞品分析", keyword.keyword])
    _append(summary, ["机会分", run.opportunity_score if run.opportunity_score is not None else "—"])
    _append(summary, ["结论", run.verdict or "数据不足"])
    _append(summary, ["数据完整度", f"{run.confidence or 0:.1f}%"])
    analysis = run.analysis or {}
    evidence = analysis.get("evidence") or {}
    collector = analysis.get("collector_health") or {}
    if evidence:
        _append(summary, ["证据等级", f"{evidence.get('grade', '—')} · {evidence.get('label', '—')}"])
        _append(summary, [
            "证据摘要",
            f"采集健康度 {evidence.get('collector_health', 0)}%；相关样本 {evidence.get('sample_total', 0)} 条；"
            + ("；".join(evidence.get("reasons") or []) or "未发现主要证据警告"),
        ])
    if collector:
        _append(summary, ["采集器状态", f"{collector.get('status', 'unknown')} · {collector.get('health_score', 0)}%"])
    _append(summary, ["数据口径", "公开搜索结果快照；不是利润、真实月销量或转化率预测。"])
    request_config = analysis.get("request_config") or {}
    if request_config:
        per_page = request_config.get("results_limit", "—")
        pages = request_config.get("search_pages", 1)
        maximum = request_config.get("max_results_per_platform")
        if maximum is None and isinstance(per_page, int) and isinstance(pages, int):
            maximum = per_page * pages
        _append(summary, [
            "本次扫描配置",
            f"平台 {', '.join(request_config.get('platforms') or [])}；前 {pages} 页；"
            f"每页最多 {per_page} 条；每平台最多 {maximum or '—'} 条",
        ])
        _append(summary, [
            "平台实际搜索词",
            request_config.get("marketplace_query") or request_config.get("keyword") or "—",
        ])
        localization = request_config.get("localization") or {}
        if localization:
            _append(summary, [
                "关键词翻译",
                f"来源 {localization.get('source', '—')}；"
                f"同义词 {', '.join(localization.get('aliases') or []) or '—'}",
            ])
    for platform, score in (run.platform_scores or {}).items():
        platform_score = score.get("score") if score.get("eligible", True) else "—"
        raw_sample = score.get("raw_sample_size", score.get("sample_size", 0))
        reasons = "；".join(score.get("eligibility_reasons") or [])
        _append(summary, [
            f"{platform.title()} 结论",
            f"{score.get('verdict', '数据不足')}；机会分 {platform_score}；相关样本 {score.get('sample_size', 0)}/{raw_sample}；完整度 {score.get('confidence', 0)}%" + (f"；{reasons}" if reasons else ""),
        ])
    for platform, health in (collector.get("platforms") or {}).items():
        warnings = "；".join(health.get("warnings") or []) or "无"
        _append(summary, [
            f"{platform.title()} 采集健康",
            f"{health.get('status', 'unknown')}；健康度 {health.get('health_score', 0)}%；"
            f"raw {health.get('raw_count', 0)} → parsed {health.get('parsed_count', 0)}；{warnings}",
        ])
    shopdora_summary = ((analysis.get("third_party") or {}).get("shopdora") or {})
    if shopdora_summary:
        metrics = shopdora_summary.get("metrics") or {}
        _append(summary, [
            "Shopdora 插件增强",
            f"覆盖 {shopdora_summary.get('sample_size', 0)}/{shopdora_summary.get('snapshot_sample_size', 0)} 条 Shopee 商品；"
            f"近30日销量中位估算 {metrics.get('median_sales_30d', '—')}；"
            "第三方估算，不参与规则机会分。",
        ])
    ai = analysis.get("ai") or {}
    if ai.get("status") == "completed":
        _append(summary, ["AI 辅助解读", ai.get("summary") or "—"])
        for finding in ai.get("findings") or []:
            _append(summary, ["AI 观察", finding])
        for risk in ai.get("risks") or []:
            _append(summary, ["AI 风险", risk])
        for action in ai.get("actions") or []:
            _append(summary, ["AI 行动", action])
        for index, step in enumerate(ai.get("next_steps") or [], 1):
            if not isinstance(step, dict):
                continue
            _append(summary, [
                f"AI 路线 {index} · {step.get('stage') or '下一步'}",
                step.get("title") or "—",
            ])
            _append(summary, ["为什么现在做", step.get("why") or "—"])
            for task in step.get("tasks") or []:
                _append(summary, ["执行动作", task])
            _append(summary, ["复盘观察项", step.get("watch") or "—"])
    for recommendation in analysis.get("recommendations", []):
        _append(summary, ["规则评分依据", recommendation])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 100

    columns = [
        "平台位次", "页码", "页内位次", "商品", "价格(MYR)", "原价(MYR)", "折扣(%)",
        "公开已售", "评分", "评论数", "店铺", "地区", "广告", "链接", "采集时间",
        "增强来源", "近30日销量估算", "近30日销量增长(%)", "近30日销售额估算(MYR)",
        "总销量估算", "GMV估算(MYR)", "插件卖家", "店铺类型", "品牌", "类目",
        "类目月销排名", "上架日期", "上架天数", "点赞数",
    ]
    for platform in ("shopee", "lazada"):
        ws = wb.create_sheet(f"{platform.title()}竞品")
        _header(ws, columns)
        for item in [row for row in rows if row.platform == platform]:
            shopdora = ((item.raw_data or {}).get("shopdora") or {})
            _append(ws, [
                item.search_rank,
                (item.raw_data or {}).get("search_page", "—"),
                (item.raw_data or {}).get("page_rank", "—"),
                item.title, item.price if item.price is not None else "—",
                item.original_price if item.original_price is not None else "—",
                item.discount_percent if item.discount_percent is not None else "—",
                item.sold_count if item.sold_count is not None else "—",
                item.rating if item.rating is not None else "—",
                item.review_count if item.review_count is not None else "—",
                item.seller_name or "—", item.seller_location or "—",
                "是" if item.is_sponsored is True else "否" if item.is_sponsored is False else "—",
                item.product_url, item.collected_at.isoformat() if item.collected_at else "—",
                "Shopdora · 第三方估算" if shopdora else "—",
                _dash(shopdora.get("sales_30d")),
                _dash(shopdora.get("sales_30d_growth_percent")),
                _dash(shopdora.get("revenue_30d_myr")),
                _dash(shopdora.get("total_sales_estimate")),
                _dash(shopdora.get("gmv_estimate_myr")),
                _dash(shopdora.get("seller_name")),
                _dash(shopdora.get("seller_type")),
                _dash(shopdora.get("brand")),
                _dash(shopdora.get("category_path")),
                _dash(shopdora.get("category_monthly_sales_rank")),
                _dash(shopdora.get("listed_at")),
                _dash(shopdora.get("listing_age_days")),
                _dash(shopdora.get("like_count")),
            ])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["N"].width = 55
        ws.column_dimensions["P"].width = 24
        ws.column_dimensions["V"].width = 28
        ws.column_dimensions["Y"].width = 48

    trend = wb.create_sheet("每日价格与排名趋势")
    _header(trend, ["采集时间", "平台", "商品ID", "商品", "价格(MYR)", "公开已售", "评论数", "平台位次", "页码", "页内位次"])
    history = (
        db.query(ListingSnapshot)
        .join(AnalysisRun, ListingSnapshot.run_id == AnalysisRun.id)
        .filter(
            ListingSnapshot.keyword_id == keyword.id,
            AnalysisRun.status.in_(RESULT_STATUSES),
        )
        .order_by(ListingSnapshot.collected_at, ListingSnapshot.id)
        .all()
    )
    for item in history:
        _append(trend, [item.collected_at.isoformat() if item.collected_at else "—", item.platform, item.item_id, item.title, item.price if item.price is not None else "—", item.sold_count if item.sold_count is not None else "—", item.review_count if item.review_count is not None else "—", item.search_rank, (item.raw_data or {}).get("search_page", "—"), (item.raw_data or {}).get("page_rank", "—")])

    notes = wb.create_sheet("数据口径说明")
    _append(notes, ["字段", "说明"])
    _append(notes, ["公开已售", "平台页面当时展示的累计/公开口径，不换算为日销量或月销量。"])
    _append(notes, ["Shopdora 插件增强", "仅在用户 Chrome 已安装并加载 Shopdora 时记录；月销量、销售额、GMV、增长等均按第三方估算字段展示，不覆盖平台公开字段，也不参与规则机会分。免费套餐可能只覆盖当前页部分商品。"])
    _append(notes, ["机会分", "需求信号40%、进入门槛35%、价格空间25%的证据门槛启发式评分。"])
    _append(notes, ["AI 辅助", "AI 只翻译严格同义词并解读已聚合的公开字段及明确标注的第三方估算，不修改机会分、证据等级或规则结论；AI 失败时规则分析仍完整可用。"])
    _append(notes, ["证据等级", "A/B/C/D 综合平台评分可用性、数据完整度、采集健康度和相关样本量；D 不输出强结论。"])
    _append(notes, ["采集健康度", "独立判断 raw 搜索卡片到可解析商品的转换、样本覆盖和关键字段覆盖，用来区分市场弱与采集器异常。"])
    _append(notes, ["缺失值", "公开页面未提供的字段不会填0，也不会把剩余权重放大；证据不足时不输出强结论。"])
    _append(notes, ["相关性", "低关键词相关性的搜索漂移结果会保留在原始快照，但从机会评分样本中排除。"])
    _append(notes, ["趋势", "仅纳入 completed/partial 任务；运行中、验证码暂停与失败任务的恢复 checkpoint 不进入趋势。"])
    _append(notes, ["平台位次", "按当页 DOM 页内位次与页面容量换算；同时保留页码和页内位次，页面未完整采集时任务标记为部分完成。"])
    _append(notes, ["跨平台", "所选平台分别评分；原始已售数字不直接跨平台比较。"])
    _append(notes, ["单元格安全", "用户输入与平台公开文本统一按文本写入；以 =、+、-、@ 开头的字符串不会作为 Excel 公式执行。"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
