import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401
from app.core.database import Base
from app.models.marketplace import AnalysisRun, ListingSnapshot, TrackedKeyword
from app.models.user import User
from app.services.marketplace.report import _excel_safe, build_report


class ReportFormulaSafetyTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine)

    def tearDown(self):
        self.engine.dispose()

    def test_excel_safe_neutralizes_formula_prefixes(self):
        self.assertEqual(_excel_safe("=1+1"), "'=1+1")
        self.assertEqual(_excel_safe(" +SUM(A1:A2)"), "' +SUM(A1:A2)")
        self.assertEqual(_excel_safe("@SUM(A1:A2)"), "'@SUM(A1:A2)")
        self.assertEqual(_excel_safe("-1+2"), "'-1+2")
        self.assertEqual(_excel_safe("ordinary text"), "ordinary text")
        self.assertEqual(_excel_safe(12.5), 12.5)

    def test_untrusted_keyword_listing_and_recommendation_are_text_cells(self):
        db = self.Session()
        user = User(username="report-user", email="report@example.com", password_hash="x")
        db.add(user)
        db.flush()
        keyword = TrackedKeyword(
            user_id=user.id,
            keyword='=HYPERLINK("https://example.test","click")',
            platforms=["shopee"],
        )
        db.add(keyword)
        db.flush()
        run = AnalysisRun(
            keyword_id=keyword.id,
            status="completed",
            opportunity_score=60,
            verdict="谨慎观察",
            confidence=80,
            platform_scores={"shopee": {"score": 60, "verdict": "谨慎观察", "sample_size": 1, "confidence": 80}},
            analysis={
                "recommendations": ['=WEBSERVICE("https://example.test")'],
                "request_config": {"platforms": ["shopee"], "results_limit": 20},
            },
        )
        db.add(run)
        db.flush()
        db.add(ListingSnapshot(
            run_id=run.id,
            keyword_id=keyword.id,
            platform="shopee",
            item_id="evil-1",
            title='=HYPERLINK("https://evil.test","product")',
            product_url="https://shopee.com.my/product/evil-1",
            price=10,
            seller_name="+SUM(A1:A9)",
            seller_location="@malicious",
            search_rank=1,
            data_quality=1.0,
        ))
        db.commit()
        db.refresh(run)

        workbook = load_workbook(BytesIO(build_report(db, run).getvalue()), data_only=False)
        summary = workbook["综合结论"]
        self.assertTrue(str(summary["B1"].value).startswith("'="))
        self.assertNotEqual(summary["B1"].data_type, "f")
        recommendation_cells = [cell.value for row in summary.iter_rows() for cell in row if isinstance(cell.value, str)]
        self.assertTrue(any(value.startswith("'=WEBSERVICE") for value in recommendation_cells))

        shopee = workbook["Shopee竞品"]
        self.assertTrue(str(shopee["B2"].value).startswith("'="))
        self.assertNotEqual(shopee["B2"].data_type, "f")
        self.assertTrue(str(shopee["I2"].value).startswith("'+"))
        self.assertTrue(str(shopee["J2"].value).startswith("'@"))

        trend = workbook["每日价格与排名趋势"]
        self.assertTrue(str(trend["D2"].value).startswith("'="))
        self.assertNotEqual(trend["D2"].data_type, "f")
        db.close()


if __name__ == "__main__":
    unittest.main()
