import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401
from app.core.database import Base
from app.models.marketplace import AnalysisRun, ListingSnapshot, TrackedKeyword
from app.models.user import User
from app.services.marketplace.adapters import MarketplaceListing
from app.services.marketplace.report import build_report
from app.services.marketplace import runner


class SnapshotAndReportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine)
        db = self.Session()
        user = User(username="tester", email="tester@example.com", password_hash="x")
        db.add(user)
        db.flush()
        keyword = TrackedKeyword(user_id=user.id, keyword="water bottle", platforms=["shopee", "lazada"])
        db.add(keyword)
        db.flush()
        run1 = AnalysisRun(keyword_id=keyword.id, status="completed", opportunity_score=61.2, verdict="谨慎观察", confidence=80)
        run2 = AnalysisRun(keyword_id=keyword.id, status="partial")
        db.add_all([run1, run2])
        db.commit()
        self.keyword_id, self.run1_id, self.run2_id = keyword.id, run1.id, run2.id
        db.close()

    def tearDown(self):
        self.engine.dispose()

    def test_later_empty_run_does_not_overwrite_previous_success(self):
        listing = MarketplaceListing(
            platform="shopee", item_id="item-1", title="Bottle", product_url="https://example.test/1",
            price=19.9, sold_count=120, search_rank=1,
        )
        db = self.Session()
        keyword = db.query(TrackedKeyword).filter_by(id=self.keyword_id).one()
        run1 = db.query(AnalysisRun).filter_by(id=self.run1_id).one()
        run2 = db.query(AnalysisRun).filter_by(id=self.run2_id).one()
        runner._persist_results(db, run1, keyword, {"shopee": [listing], "lazada": []})
        db.commit()
        runner._persist_results(db, run2, keyword, {"shopee": [], "lazada": []})
        db.commit()
        self.assertEqual(db.query(ListingSnapshot).filter_by(run_id=self.run1_id).count(), 1)
        self.assertEqual(db.query(ListingSnapshot).filter_by(run_id=self.run2_id).count(), 0)
        db.close()

    def test_excel_has_required_sheets_and_dash_for_missing_values(self):
        db = self.Session()
        run = db.query(AnalysisRun).filter_by(id=self.run1_id).one()
        run.platform_scores = {
            "shopee": {"score": 62, "verdict": "谨慎观察", "sample_size": 12, "confidence": 81},
            "lazada": {"score": 60, "verdict": "谨慎观察", "sample_size": 11, "confidence": 79},
        }
        run.analysis = {
            "request_config": {
                "keyword": "water bottle",
                "marketplace_query": "water bottle",
                "platforms": ["shopee", "lazada"],
                "results_limit": 20,
                "search_pages": 3,
                "max_results_per_platform": 60,
            },
            "recommendations": ["先小批量测试。"],
            "third_party": {"shopdora": {
                "sample_size": 1,
                "snapshot_sample_size": 1,
                "metrics": {"median_sales_30d": 916},
            }},
            "ai": {
                "status": "completed",
                "summary": "先验证公开信号。",
                "findings": ["价格带值得复核。"],
                "risks": ["真实成本未知。"],
                "actions": ["核验供应链。"],
                "next_steps": [{
                    "stage": "先核验",
                    "title": "复核目标商品",
                    "why": "公开数据只能支持继续验证。",
                    "tasks": ["记录采购成本。", "核对商品规格。"],
                    "watch": "复盘真实成本和规格是否可比。",
                }],
            },
        }
        db.add(ListingSnapshot(
            run_id=run.id, keyword_id=self.keyword_id, platform="lazada", item_id="l1",
            title="Bottle", product_url="https://example.test/l1", price=20, search_rank=24,
            data_quality=0.5, raw_data={"search_page": 2, "page_rank": 4, "page_size": 20},
        ))
        db.add(ListingSnapshot(
            run_id=run.id, keyword_id=self.keyword_id, platform="shopee", item_id="s1",
            title="Towel", product_url="https://example.test/s1", price=2.8, search_rank=1,
            raw_data={
                "search_page": 1,
                "page_rank": 1,
                "page_size": 20,
                "shopdora": {
                    "provider": "Shopdora", "estimated": True,
                    "sales_30d": 916, "sales_30d_growth_percent": 33.28,
                    "revenue_30d_myr": 2564.8, "total_sales_estimate": 27380,
                    "gmv_estimate_myr": 76664.0, "seller_name": "Simplico",
                    "seller_type": "本土", "brand": None, "category_path": "家居-毛巾",
                    "category_monthly_sales_rank": 48, "listed_at": "2022-08-15",
                    "listing_age_days": 1470, "like_count": 484,
                },
            },
        ))
        db.commit()
        output = build_report(db, run)
        workbook = load_workbook(BytesIO(output.getvalue()))
        self.assertEqual(
            workbook.sheetnames,
            ["综合结论", "Shopee竞品", "Lazada竞品", "每日价格与排名趋势", "数据口径说明"],
        )
        lazada = workbook["Lazada竞品"]
        self.assertEqual(lazada.cell(2, 1).value, 24)
        self.assertEqual(lazada.cell(2, 2).value, 2)
        self.assertEqual(lazada.cell(2, 3).value, 4)
        self.assertEqual(lazada.cell(2, 6).value, "—")
        summary_rows = list(workbook["综合结论"].iter_rows())
        summary_labels = [row[0].value for row in summary_rows]
        self.assertIn("Lazada 结论", summary_labels)
        self.assertIn("本次扫描配置", summary_labels)
        query_row = next(row for row in summary_rows if row[0].value == "平台实际搜索词")
        self.assertEqual(query_row[1].value, "water bottle")
        self.assertIn("AI 路线 1 · 先核验", summary_labels)
        self.assertIn("复盘观察项", summary_labels)
        self.assertIn("规则评分依据", summary_labels)
        self.assertIn("Shopdora 插件增强", summary_labels)
        shopee = workbook["Shopee竞品"]
        self.assertEqual(shopee.cell(2, 16).value, "Shopdora · 第三方估算")
        self.assertEqual(shopee.cell(2, 17).value, 916)
        self.assertEqual(shopee.cell(2, 19).value, 2564.8)
        self.assertEqual(shopee.cell(2, 25).value, "家居-毛巾")
        notes = [row[0].value for row in workbook["数据口径说明"].iter_rows(min_row=2)]
        self.assertIn("Shopdora 插件增强", notes)
        trend = workbook["每日价格与排名趋势"]
        self.assertEqual(trend.cell(2, 8).value, 24)
        self.assertEqual(trend.cell(2, 9).value, 2)
        self.assertEqual(trend.cell(2, 10).value, 4)
        db.close()

    def test_trend_excludes_running_verification_and_failed_checkpoints(self):
        db = self.Session()
        stable = db.query(AnalysisRun).filter_by(id=self.run1_id).one()
        stable.analysis = {"recommendations": []}
        db.add(ListingSnapshot(
            run_id=stable.id, keyword_id=self.keyword_id, platform="shopee", item_id="stable",
            title="Stable Bottle", product_url="https://example.test/stable", price=20, search_rank=1,
        ))
        unfinished_runs = [
            AnalysisRun(keyword_id=self.keyword_id, status="running"),
            AnalysisRun(keyword_id=self.keyword_id, status="needs_verification"),
            AnalysisRun(keyword_id=self.keyword_id, status="failed"),
        ]
        db.add_all(unfinished_runs)
        db.flush()
        for index, unfinished in enumerate(unfinished_runs):
            db.add(ListingSnapshot(
                run_id=unfinished.id, keyword_id=self.keyword_id, platform="shopee", item_id=f"temp-{index}",
                title=f"Temp {index}", product_url=f"https://example.test/temp-{index}", price=10 + index, search_rank=1,
            ))
        db.commit()

        workbook = load_workbook(BytesIO(build_report(db, stable).getvalue()))
        trend = workbook["每日价格与排名趋势"]
        item_ids = [row[2].value for row in trend.iter_rows(min_row=2)]
        self.assertIn("stable", item_ids)
        self.assertNotIn("temp-0", item_ids)
        self.assertNotIn("temp-1", item_ids)
        self.assertNotIn("temp-2", item_ids)
        db.close()


if __name__ == "__main__":
    unittest.main()
